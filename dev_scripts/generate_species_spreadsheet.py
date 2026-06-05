#!/usr/bin/env python3
"""Phase 1 preprocessor for species comparison extraction.

This script only implements preprocessing and filtering.
It flattens C preprocessor branches for common project conditions,
collects active macros, and removes unwanted species blocks:
- Gigantamax forms
- Cosmetic-only forms (explicit family allowlist)

It also normalizes embedded symbol markers like '#sym:VICTREEBEL_SP_DEF'
by treating them as 0 for now.
"""

from __future__ import annotations

import argparse
import ast
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


GEN_FILES = [f"gen_{i}_families.h" for i in range(1, 10)]
SPECIES_INFO_SUBPATH = Path("src/data/pokemon/species_info")


COSMETIC_SPECIES_PREFIXES = (
    "SPECIES_ALCREMIE_",
    "SPECIES_UNOWN_",
    "SPECIES_TATSUGIRI_",
    "SPECIES_FLOETTE_",
    "SPECIES_FLABEBE_",
    "SPECIES_FLORGES_",
    "SPECIES_MINIOR_",
    "SPECIES_VIVILLON_",
    "SPECIES_SPEWPA_",
    "SPECIES_SCATTERBUG_",
)


COSMETIC_SPECIES_EXACT = {
    # Keep one canonical entry for each cosmetic family.
    # Base form rows stay; alternates are removed by prefix checks above.
}


DIRECTIVE_IF_RE = re.compile(r"^\s*#\s*if\b(.*)$")
DIRECTIVE_IFDEF_RE = re.compile(r"^\s*#\s*ifdef\b(.*)$")
DIRECTIVE_IFNDEF_RE = re.compile(r"^\s*#\s*ifndef\b(.*)$")
DIRECTIVE_ELIF_RE = re.compile(r"^\s*#\s*elif\b(.*)$")
DIRECTIVE_ELSE_RE = re.compile(r"^\s*#\s*else\b")
DIRECTIVE_ENDIF_RE = re.compile(r"^\s*#\s*endif\b")
DIRECTIVE_DEFINE_RE = re.compile(r"^\s*#\s*define\s+([A-Za-z_][A-Za-z0-9_:]*)\s*(.*)$")
DIRECTIVE_DEFINE_EMBEDDED_SYM_RE = re.compile(r"^\s*#\s*define\s+(#sym:[A-Za-z_][A-Za-z0-9_]*)\s*(.*)$")
SPECIES_ASSIGN_RE = re.compile(r"\[\s*(SPECIES_[A-Z0-9_]+)\s*\]\s*=")
EMBEDDED_SYM_RE = re.compile(r"#sym:[A-Za-z_][A-Za-z0-9_]*")
STAT_FIELD_RE = re.compile(r"^\s*\.(baseHP|baseAttack|baseDefense|baseSpeed|baseSpAttack|baseSpDefense)\s*=\s*(.+?),\s*$")
TYPES_FIELD_RE = re.compile(r"^\s*\.types\s*=\s*MON_TYPES\((.+)\),\s*$")
ABILITIES_FIELD_RE = re.compile(r"^\s*\.abilities\s*=\s*\{(.+)\},\s*$")
SPECIES_NAME_RE = re.compile(r'^\s*\.speciesName\s*=\s*_\("([^"]+)"\),\s*$')
TERNARY_PAREN_RE = re.compile(r"\(([^()?:]+?)\?([^:()]+):([^()]+)\)")
TERNARY_FLAT_RE = re.compile(r"^\s*([^?:]+?)\?([^:]+):(.+)$")
MACRO_TOKEN_RE = re.compile(r"\b([A-Z_][A-Z0-9_]*)\b")
IDENTIFIER_RE = re.compile(r"\b([A-Za-z_][A-Za-z0-9_]*)\b")
ABILITIES_MACRO_FIELD_RE = re.compile(r"^\s*\.abilities\s*=\s*([A-Z_][A-Z0-9_]*)\s*,\s*$")

DEFAULT_CLEAN_REPO_ROOT = Path(r"C:\Users\nk619\Documents\pokeemerald-expansion-1.15.0 clean\pokeemerald-expansion-master")
DEFAULT_OUTPUT_FILE = "species_comparison.xlsx"

HEADER_TITLES = [
    "Species",
    "Version",
    "Type 1",
    "Type 2",
    "Ability 1",
    "Ability 2",
    "Hidden Ability",
    "HP",
    "Attack",
    "Defense",
    "Sp. Atk",
    "Sp. Def",
    "Speed",
    "BST",
]

STAT_COLUMNS = {
    "H": "00CCFFCC",  # HP
    "I": "00FFFACC",  # Attack
    "J": "00FFE0C0",  # Defense
    "K": "00CCE5FF",  # Sp. Atk
    "L": "00AACCDD",  # Sp. Def
    "M": "00FFCCFF",  # Speed
    "N": "00EEEEEE",  # BST
}

TYPE_COLORS = {
    "Normal": "00C8C8A0",
    "Fire": "00FFAA66",
    "Water": "0099CCFF",
    "Grass": "0099DD88",
    "Electric": "00FFDD55",
    "Ice": "00BBEEEE",
    "Fighting": "00CC9966",
    "Poison": "00CC88CC",
    "Ground": "00EEDD99",
    "Flying": "00AABBEE",
    "Psychic": "00FFAAAA",
    "Bug": "00AABB77",
    "Rock": "00BBBB99",
    "Ghost": "009977AA",
    "Dragon": "0099AADD",
    "Dark": "00997766",
    "Steel": "00BBBBCC",
    "Fairy": "00FFBBDD",
    "Stellar": "00DDBB77",
}

THIN_SIDE = Side(style="thin", color="000000")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
HEADER_FILL = PatternFill(fill_type="solid", fgColor="00DDDDDD")
FORM_TOKEN_LABELS = {
    "ALOLA": "Alola",
    "GALAR": "Galar",
    "HISUI": "Hisui",
    "PALDEA": "Paldea",
    "MEGA": "Mega",
    "PRIMAL": "Primal",
    "THERIAN": "Therian",
    "INCARNATE": "Incarnate",
    "ORIGIN": "Origin",
    "RESOLUTE": "Resolute",
    "COMPLETE": "Complete",
    "ZEN": "Zen",
    "CROWNED": "Crowned",
    "BLOODMOON": "Bloodmoon",
    "WHITE": "White",
    "BLACK": "Black",
    "X": "X",
    "Y": "Y",
    "Z": "Z",
}


@dataclass
class SpeciesRecord:
    species_key: str
    species_name: str
    type1: str
    type2: str
    ability1: str
    ability2: str
    hidden_ability: str
    base_hp: int
    base_attack: int
    base_defense: int
    base_sp_attack: int
    base_sp_defense: int
    base_speed: int


@dataclass
class MergedSpeciesRecord:
    species_key: str
    species_name: str
    original: Optional[SpeciesRecord]
    customized: Optional[SpeciesRecord]


@dataclass
class ConditionalFrame:
    parent_active: bool
    branch_taken: bool
    current_active: bool
    gmax_frame: bool


@dataclass
class PreprocessResult:
    lines: List[str]
    macros: Dict[str, str]
    excluded_gmax: int
    excluded_cosmetic: int


class Phase1Preprocessor:
    def __init__(self) -> None:
        self.macros: Dict[str, str] = {}

    def preprocess_file(self, file_path: Path) -> PreprocessResult:
        self.macros = {}
        raw_lines = file_path.read_text(encoding="utf-8").splitlines()
        stack: List[ConditionalFrame] = []
        gmax_depth = 0
        active_lines_with_flags: List[Tuple[str, bool]] = []

        for line in raw_lines:
            if m := DIRECTIVE_IF_RE.match(line):
                cond = m.group(1).strip()
                cond_value = self._evaluate_condition(cond)
                parent_active = self._is_currently_active(stack)
                current_active = parent_active and cond_value
                gmax_frame = "P_GIGANTAMAX_FORMS" in cond
                if gmax_frame and current_active:
                    gmax_depth += 1
                stack.append(
                    ConditionalFrame(
                        parent_active=parent_active,
                        branch_taken=current_active,
                        current_active=current_active,
                        gmax_frame=gmax_frame,
                    )
                )
                continue

            if m := DIRECTIVE_IFDEF_RE.match(line):
                cond = m.group(1).strip()
                cond_value = self._evaluate_condition(cond)
                parent_active = self._is_currently_active(stack)
                current_active = parent_active and cond_value
                gmax_frame = "P_GIGANTAMAX_FORMS" in cond
                if gmax_frame and current_active:
                    gmax_depth += 1
                stack.append(
                    ConditionalFrame(
                        parent_active=parent_active,
                        branch_taken=current_active,
                        current_active=current_active,
                        gmax_frame=gmax_frame,
                    )
                )
                continue

            if m := DIRECTIVE_IFNDEF_RE.match(line):
                cond = m.group(1).strip()
                cond_value = not self._evaluate_condition(cond)
                parent_active = self._is_currently_active(stack)
                current_active = parent_active and cond_value
                gmax_frame = "P_GIGANTAMAX_FORMS" in cond
                if gmax_frame and current_active:
                    gmax_depth += 1
                stack.append(
                    ConditionalFrame(
                        parent_active=parent_active,
                        branch_taken=current_active,
                        current_active=current_active,
                        gmax_frame=gmax_frame,
                    )
                )
                continue

            if m := DIRECTIVE_ELIF_RE.match(line):
                cond = m.group(1).strip()
                if not stack:
                    continue
                frame = stack[-1]
                if frame.gmax_frame and frame.current_active:
                    gmax_depth -= 1
                if frame.parent_active and not frame.branch_taken and self._evaluate_condition(cond):
                    frame.current_active = True
                    frame.branch_taken = True
                else:
                    frame.current_active = False
                if frame.gmax_frame and frame.current_active:
                    gmax_depth += 1
                continue

            if DIRECTIVE_ELSE_RE.match(line):
                if not stack:
                    continue
                frame = stack[-1]
                if frame.gmax_frame and frame.current_active:
                    gmax_depth -= 1
                frame.current_active = frame.parent_active and not frame.branch_taken
                frame.branch_taken = True
                if frame.gmax_frame and frame.current_active:
                    gmax_depth += 1
                continue

            if DIRECTIVE_ENDIF_RE.match(line):
                if stack:
                    frame = stack.pop()
                    if frame.gmax_frame and frame.current_active:
                        gmax_depth -= 1
                continue

            if not self._is_currently_active(stack):
                continue

            if m := DIRECTIVE_DEFINE_EMBEDDED_SYM_RE.match(line):
                macro_name = m.group(1)
                self.macros[macro_name] = "0"
                self.macros[macro_name.split(":", maxsplit=1)[1]] = "0"
                continue

            if m := DIRECTIVE_DEFINE_RE.match(line):
                macro_name = m.group(1)
                macro_value = m.group(2).strip() or "1"
                macro_value = EMBEDDED_SYM_RE.sub("0", macro_value)

                if macro_name.startswith("#sym:"):
                    # Embedded symbols are placeholders; treat them as 0 for now.
                    self.macros[macro_name] = "0"
                    self.macros[macro_name.split(":", maxsplit=1)[1]] = "0"
                else:
                    self.macros[macro_name] = macro_value
                continue

            normalized = EMBEDDED_SYM_RE.sub("0", line)
            active_lines_with_flags.append((normalized, gmax_depth > 0))

        filtered_lines, excluded_gmax, excluded_cosmetic = self._filter_species_blocks(active_lines_with_flags)
        return PreprocessResult(
            lines=filtered_lines,
            macros=dict(self.macros),
            excluded_gmax=excluded_gmax,
            excluded_cosmetic=excluded_cosmetic,
        )

    @staticmethod
    def _is_currently_active(stack: List[ConditionalFrame]) -> bool:
        return all(frame.current_active for frame in stack)

    @staticmethod
    def _evaluate_condition(condition: str) -> bool:
        cond = condition.strip()
        if not cond:
            return False

        # Force these to False so the most common data branches match current usage goals.
        if "P_GBA_STYLE_SPECIES_GFX" in cond and "!P_GBA_STYLE_SPECIES_GFX" not in cond:
            return False
        if "P_CUSTOM_GENDER_DIFF_ICONS" in cond and "!P_CUSTOM_GENDER_DIFF_ICONS" not in cond:
            return False

        # Everything else is treated as enabled for fast, practical flattening.
        return True

    def _filter_species_blocks(
        self, lines_with_flags: List[Tuple[str, bool]]
    ) -> Tuple[List[str], int, int]:
        out_lines: List[str] = []
        i = 0
        excluded_gmax = 0
        excluded_cosmetic = 0

        while i < len(lines_with_flags):
            line, in_gmax_scope = lines_with_flags[i]
            start_match = SPECIES_ASSIGN_RE.search(line)
            if not start_match:
                out_lines.append(line)
                i += 1
                continue

            species_key = start_match.group(1)
            block_lines = [line]
            block_gmax = in_gmax_scope
            brace_depth = line.count("{") - line.count("}")
            i += 1

            # Common format is split over lines:
            # [SPECIES_X] =
            # {
            while i < len(lines_with_flags) and brace_depth == 0:
                probe_line, probe_flag = lines_with_flags[i]
                block_lines.append(probe_line)
                block_gmax = block_gmax or probe_flag
                brace_depth += probe_line.count("{") - probe_line.count("}")
                i += 1
                if brace_depth > 0:
                    break

            while i < len(lines_with_flags):
                block_line, block_flag = lines_with_flags[i]
                block_lines.append(block_line)
                block_gmax = block_gmax or block_flag
                brace_depth += block_line.count("{") - block_line.count("}")
                i += 1
                if brace_depth <= 0:
                    break

            if block_gmax or species_key.endswith("_GMAX"):
                excluded_gmax += 1
                continue

            if self._is_cosmetic_species(species_key):
                excluded_cosmetic += 1
                continue

            out_lines.extend(block_lines)

        return out_lines, excluded_gmax, excluded_cosmetic

    @staticmethod
    def _is_cosmetic_species(species_key: str) -> bool:
        if species_key in COSMETIC_SPECIES_EXACT:
            return True
        return species_key.startswith(COSMETIC_SPECIES_PREFIXES)


class Phase2SpeciesParser:
    def parse_species(self, preprocess_result: PreprocessResult) -> List[SpeciesRecord]:
        lines = preprocess_result.lines
        macros = preprocess_result.macros
        records: List[SpeciesRecord] = []
        i = 0

        while i < len(lines):
            start_match = SPECIES_ASSIGN_RE.search(lines[i])
            if not start_match:
                i += 1
                continue

            species_key = start_match.group(1)
            block_lines = [lines[i]]
            brace_depth = lines[i].count("{") - lines[i].count("}")
            i += 1

            while i < len(lines) and brace_depth == 0:
                block_lines.append(lines[i])
                brace_depth += lines[i].count("{") - lines[i].count("}")
                i += 1
                if brace_depth > 0:
                    break

            while i < len(lines):
                block_lines.append(lines[i])
                brace_depth += lines[i].count("{") - lines[i].count("}")
                i += 1
                if brace_depth <= 0:
                    break

            record = self._parse_species_block(species_key, block_lines, macros)
            if record is not None:
                records.append(record)

        return records

    def _parse_species_block(
        self,
        species_key: str,
        block_lines: List[str],
        macros: Dict[str, str],
    ) -> Optional[SpeciesRecord]:
        values: Dict[str, int] = {}
        species_name = species_key.replace("SPECIES_", "").replace("_", " ").title()
        type1 = "TYPE_NONE"
        type2 = "TYPE_NONE"
        ability1 = "ABILITY_NONE"
        ability2 = "ABILITY_NONE"
        hidden_ability = "ABILITY_NONE"

        for line in block_lines:
            line_no_comment = strip_trailing_line_comment(line)

            if m := SPECIES_NAME_RE.match(line_no_comment):
                species_name = m.group(1)
                continue

            if m := TYPES_FIELD_RE.match(line_no_comment):
                parsed_types = [part.strip() for part in m.group(1).split(",") if part.strip()]
                if parsed_types:
                    type1 = parsed_types[0]
                if len(parsed_types) >= 2:
                    type2 = parsed_types[1]
                continue

            if m := ABILITIES_FIELD_RE.match(line_no_comment):
                parsed_abilities = [part.strip() for part in m.group(1).split(",") if part.strip()]
                if parsed_abilities:
                    ability1 = parsed_abilities[0]
                if len(parsed_abilities) >= 2:
                    ability2 = parsed_abilities[1]
                if len(parsed_abilities) >= 3:
                    hidden_ability = parsed_abilities[2]
                continue

            if m := ABILITIES_MACRO_FIELD_RE.match(line_no_comment):
                macro_name = m.group(1)
                macro_value = macros.get(macro_name, "")
                if macro_value.startswith("{") and macro_value.endswith("}"):
                    parsed_abilities = [part.strip() for part in macro_value.strip("{} ").split(",") if part.strip()]
                    if parsed_abilities:
                        ability1 = parsed_abilities[0]
                    if len(parsed_abilities) >= 2:
                        ability2 = parsed_abilities[1]
                    if len(parsed_abilities) >= 3:
                        hidden_ability = parsed_abilities[2]
                continue

            if m := STAT_FIELD_RE.match(line_no_comment):
                field_name = m.group(1)
                expression = m.group(2).strip()
                values[field_name] = Phase3ExpressionEvaluator.eval_stat(expression, macros)

        required_fields = {
            "baseHP",
            "baseAttack",
            "baseDefense",
            "baseSpAttack",
            "baseSpDefense",
            "baseSpeed",
        }
        if not required_fields.issubset(values):
            return None

        species_name = build_display_species_name(species_key, species_name)

        return SpeciesRecord(
            species_key=species_key,
            species_name=species_name,
            type1=type1,
            type2=type2,
            ability1=ability1,
            ability2=ability2,
            hidden_ability=hidden_ability,
            base_hp=values["baseHP"],
            base_attack=values["baseAttack"],
            base_defense=values["baseDefense"],
            base_sp_attack=values["baseSpAttack"],
            base_sp_defense=values["baseSpDefense"],
            base_speed=values["baseSpeed"],
        )


class _SafeArithmeticEvaluator(ast.NodeVisitor):
    def visit_Expression(self, node: ast.Expression) -> int:
        return self.visit(node.body)

    def visit_Constant(self, node: ast.Constant) -> int:
        if isinstance(node.value, (int, float)):
            return int(node.value)
        raise ValueError("Unsupported constant in expression")

    def visit_UnaryOp(self, node: ast.UnaryOp) -> int:
        value = self.visit(node.operand)
        if isinstance(node.op, ast.UAdd):
            return value
        if isinstance(node.op, ast.USub):
            return -value
        raise ValueError("Unsupported unary operation")

    def visit_BinOp(self, node: ast.BinOp) -> int:
        left = self.visit(node.left)
        right = self.visit(node.right)

        if isinstance(node.op, ast.Add):
            return left + right
        if isinstance(node.op, ast.Sub):
            return left - right
        if isinstance(node.op, ast.Mult):
            return left * right
        if isinstance(node.op, ast.Div):
            return int(left / right)
        if isinstance(node.op, ast.FloorDiv):
            return left // right
        if isinstance(node.op, ast.Mod):
            return left % right
        if isinstance(node.op, ast.LShift):
            return left << right
        if isinstance(node.op, ast.RShift):
            return left >> right
        if isinstance(node.op, ast.BitAnd):
            return left & right
        if isinstance(node.op, ast.BitOr):
            return left | right
        if isinstance(node.op, ast.BitXor):
            return left ^ right
        raise ValueError("Unsupported binary operation")

    def generic_visit(self, node: ast.AST) -> int:
        raise ValueError(f"Unsupported expression node: {type(node).__name__}")


class Phase3ExpressionEvaluator:
    @classmethod
    def eval_stat(cls, expression: str, macros: Dict[str, str]) -> int:
        expr = cls._strip_comments(expression)
        expr = EMBEDDED_SYM_RE.sub("0", expr)
        expr = cls._expand_macros(expr, macros)
        expr = cls._resolve_ternaries(expr)
        expr = cls._normalize_identifiers(expr)
        expr = expr.strip()

        if not expr:
            return 0

        try:
            parsed = ast.parse(expr, mode="eval")
            evaluator = _SafeArithmeticEvaluator()
            return int(evaluator.visit(parsed))
        except Exception:
            return 0

    @staticmethod
    def _strip_comments(expression: str) -> str:
        return re.sub(r"/\*.*?\*/", "", expression)

    @classmethod
    def _expand_macros(cls, expression: str, macros: Dict[str, str]) -> str:
        expr = expression

        for _ in range(10):
            replaced_any = False

            def repl(match: re.Match[str]) -> str:
                nonlocal replaced_any
                token = match.group(1)
                if token not in macros:
                    return token
                macro_value = macros[token].strip()
                if "{" in macro_value or "}" in macro_value:
                    return token
                replaced_any = True
                return f"({macro_value})"

            new_expr = MACRO_TOKEN_RE.sub(repl, expr)
            expr = new_expr
            if not replaced_any:
                break

        return expr

    @classmethod
    def _resolve_ternaries(cls, expression: str) -> str:
        expr = expression

        for _ in range(20):
            updated = False

            def paren_repl(match: re.Match[str]) -> str:
                nonlocal updated
                updated = True
                true_value = match.group(2).strip()
                return f"({true_value})"

            new_expr = TERNARY_PAREN_RE.sub(paren_repl, expr)
            expr = new_expr

            if not updated and "?" in expr and ":" in expr:
                flat_match = TERNARY_FLAT_RE.match(expr)
                if flat_match:
                    updated = True
                    expr = flat_match.group(2).strip()

            if not updated:
                break

        return expr

    @staticmethod
    def _normalize_identifiers(expression: str) -> str:
        def repl(match: re.Match[str]) -> str:
            token = match.group(1)
            if token in {"TRUE", "true"}:
                return "1"
            if token in {"FALSE", "false"}:
                return "0"
            if token.startswith("0x") or token.startswith("0X"):
                return token
            if token.isdigit():
                return token
            return "0"

        return IDENTIFIER_RE.sub(repl, expression)


def preprocess_repo(repo_root: Path) -> Dict[str, PreprocessResult]:
    preprocessor = Phase1Preprocessor()
    results: Dict[str, PreprocessResult] = {}

    for gen_file in GEN_FILES:
        file_path = repo_root / SPECIES_INFO_SUBPATH / gen_file
        if not file_path.exists():
            continue
        results[gen_file] = preprocessor.preprocess_file(file_path)

    return results


def parse_repo_species(repo_root: Path) -> Dict[str, List[SpeciesRecord]]:
    preprocess_results = preprocess_repo(repo_root)
    parser = Phase2SpeciesParser()
    parsed: Dict[str, List[SpeciesRecord]] = {}

    for gen_file, result in preprocess_results.items():
        parsed[gen_file] = parser.parse_species(result)

    return parsed


def merge_species(
    original: List[SpeciesRecord],
    customized: List[SpeciesRecord],
) -> List[MergedSpeciesRecord]:
    merged: List[MergedSpeciesRecord] = []
    original_by_key = {record.species_key: record for record in original}
    custom_by_key = {record.species_key: record for record in customized}
    seen_keys = set()

    for custom in customized:
        orig = original_by_key.get(custom.species_key)
        merged.append(
            MergedSpeciesRecord(
                species_key=custom.species_key,
                species_name=custom.species_name,
                original=orig,
                customized=custom,
            )
        )
        seen_keys.add(custom.species_key)

    for orig in original:
        if orig.species_key in seen_keys:
            continue
        merged.append(
            MergedSpeciesRecord(
                species_key=orig.species_key,
                species_name=orig.species_name,
                original=orig,
                customized=custom_by_key.get(orig.species_key),
            )
        )

    return merged


def strip_trailing_line_comment(line: str) -> str:
    comment_index = line.find("//")
    if comment_index == -1:
        return line
    return line[:comment_index].rstrip()


def normalize_name_for_species_key(species_name: str) -> str:
    normalized = species_name.upper()
    normalized = normalized.replace(".", "")
    normalized = normalized.replace("'", "")
    normalized = normalized.replace("-", "_")
    normalized = normalized.replace(" ", "_")
    normalized = normalized.replace("%", "")
    normalized = re.sub(r"_+", "_", normalized).strip("_")
    return normalized


def format_form_suffix(suffix: str) -> str:
    tokens = [token for token in suffix.split("_") if token]
    if not tokens:
        return ""
    return " ".join(FORM_TOKEN_LABELS.get(token, token.title()) for token in tokens)


def build_display_species_name(species_key: str, species_name: str) -> str:
    key_body = species_key.removeprefix("SPECIES_")
    normalized_name = normalize_name_for_species_key(species_name)

    if key_body == normalized_name:
        return species_name

    suffix = ""
    if normalized_name and key_body.startswith(normalized_name + "_"):
        suffix = key_body[len(normalized_name) + 1 :]
    elif "_MEGA_" in key_body:
        suffix = key_body.split("_MEGA_", maxsplit=1)[1]
        suffix = f"MEGA_{suffix}"
    elif key_body.endswith("_MEGA"):
        suffix = "MEGA"
    elif key_body.endswith("_ALOLA"):
        suffix = "ALOLA"
    elif key_body.endswith("_GALAR"):
        suffix = "GALAR"
    elif key_body.endswith("_HISUI"):
        suffix = "HISUI"
    elif key_body.endswith("_PALDEA"):
        suffix = "PALDEA"

    formatted_suffix = format_form_suffix(suffix)
    if not formatted_suffix:
        return species_name
    return f"{species_name} {formatted_suffix}"


def format_type(type_value: str) -> str:
    if not type_value or type_value == "TYPE_NONE":
        return ""
    if type_value.startswith("TYPE_"):
        return type_value.replace("TYPE_", "").replace("_", " ").title()
    return type_value.replace("_", " ").title()


def format_ability(ability_value: str) -> str:
    if not ability_value or ability_value == "ABILITY_NONE":
        return ""
    if ability_value.startswith("ABILITY_"):
        return ability_value.replace("ABILITY_", "").replace("_", " ").title()
    return ability_value.replace("_", " ").title()


def species_row_values(species: Optional[SpeciesRecord]) -> List[object]:
    if species is None:
        return ["", "", "", "", "", "", "", "", "", "", "", ""]

    type1 = format_type(species.type1)
    type2_raw = format_type(species.type2)
    type2 = "" if not type2_raw or type2_raw == type1 else type2_raw
    a1 = format_ability(species.ability1)
    a2 = format_ability(species.ability2)
    ha = format_ability(species.hidden_ability)
    bst = (
        species.base_hp
        + species.base_attack
        + species.base_defense
        + species.base_sp_attack
        + species.base_sp_defense
        + species.base_speed
    )

    return [
        type1,
        type2,
        a1,
        a2,
        ha,
        species.base_hp,
        species.base_attack,
        species.base_defense,
        species.base_sp_attack,
        species.base_sp_defense,
        species.base_speed,
        bst,
    ]


def apply_column_styles(ws, row_idx: int) -> None:
    for col, color in STAT_COLUMNS.items():
        cell = ws[f"{col}{row_idx}"]
        cell.fill = PatternFill(fill_type="solid", fgColor=color)

    for col in ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N"]:
        ws[f"{col}{row_idx}"].border = THIN_BORDER


def apply_type_fill(ws, row_idx: int, col: str) -> None:
    value = ws[f"{col}{row_idx}"].value
    if isinstance(value, str) and value in TYPE_COLORS:
        ws[f"{col}{row_idx}"].fill = PatternFill(fill_type="solid", fgColor=TYPE_COLORS[value])


def autosize_columns(ws) -> None:
    for col_idx in range(1, 15):
        letter = get_column_letter(col_idx)
        max_len = len(HEADER_TITLES[col_idx - 1])
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, values_only=True):
            value = row[0]
            if value is None:
                continue
            max_len = max(max_len, len(str(value)))
        ws.column_dimensions[letter].width = min(max_len + 2, 42)


def build_workbook(
    merged_by_gen: Dict[str, List[MergedSpeciesRecord]],
    output_path: Path,
) -> None:
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    for gen_file in GEN_FILES:
        gen_num = gen_file.split("_")[1]
        sheet_name = f"Gen {gen_num}"
        ws = wb.create_sheet(title=sheet_name)
        ws.freeze_panes = "A2"

        for idx, title in enumerate(HEADER_TITLES, start=1):
            cell = ws.cell(row=1, column=idx, value=title)
            cell.font = Font(bold=True)
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER

        row_idx = 2
        merged_rows = merged_by_gen.get(gen_file, [])
        for merged in merged_rows:
            original_values = species_row_values(merged.original)
            custom_values = species_row_values(merged.customized)

            ws.cell(row=row_idx, column=1, value=merged.species_name)
            ws.cell(row=row_idx, column=2, value="Original")
            ws.cell(row=row_idx + 1, column=2, value="Customized")

            for offset, value in enumerate(original_values, start=3):
                ws.cell(row=row_idx, column=offset, value=value)
            for offset, value in enumerate(custom_values, start=3):
                ws.cell(row=row_idx + 1, column=offset, value=value)

            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx + 1, end_column=1)
            ws.cell(row=row_idx, column=1).alignment = Alignment(vertical="center")

            apply_column_styles(ws, row_idx)
            apply_column_styles(ws, row_idx + 1)
            apply_type_fill(ws, row_idx, "C")
            apply_type_fill(ws, row_idx, "D")
            apply_type_fill(ws, row_idx + 1, "C")
            apply_type_fill(ws, row_idx + 1, "D")

            for col_idx in range(3, 15):
                original_cell = ws.cell(row=row_idx, column=col_idx)
                custom_cell = ws.cell(row=row_idx + 1, column=col_idx)
                if (original_cell.value or "") != (custom_cell.value or ""):
                    original_cell.font = Font(bold=True, italic=True)
                    custom_cell.font = Font(bold=True, italic=True)

            row_idx += 2

        autosize_columns(ws)

    wb.save(output_path)


def build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Species comparison spreadsheet generator")
    parser.add_argument(
        "--repo-root",
        type=Path,
        default=Path(__file__).resolve().parents[1],
        help="Root path of pokeemerald-expansion repository",
    )
    parser.add_argument(
        "--clean-repo-root",
        type=Path,
        default=DEFAULT_CLEAN_REPO_ROOT,
        help="Root path of clean pokeemerald-expansion repository",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path(__file__).resolve().parents[1] / DEFAULT_OUTPUT_FILE,
        help="Output xlsx path",
    )
    parser.add_argument(
        "--emit-dir",
        type=Path,
        default=None,
        help="Optional directory to write flattened gen files",
    )
    parser.add_argument(
        "--debug-parse",
        action="store_true",
        help="Run Phase 2 parsing and print species counts per gen",
    )
    parser.add_argument(
        "--sample-species",
        type=str,
        default="SPECIES_VICTREEBEL",
        help="Species key to print after parsing in debug mode",
    )
    return parser


def main() -> None:
    args = build_arg_parser().parse_args()
    repo_root = args.repo_root
    clean_repo_root = args.clean_repo_root
    output_path = args.output

    results = preprocess_repo(repo_root)

    if args.emit_dir is not None:
        args.emit_dir.mkdir(parents=True, exist_ok=True)

    total_gmax = 0
    total_cosmetic = 0

    for gen_file, result in sorted(results.items()):
        total_gmax += result.excluded_gmax
        total_cosmetic += result.excluded_cosmetic

        if args.emit_dir is not None:
            out_file = args.emit_dir / gen_file
            out_file.write_text("\n".join(result.lines) + "\n", encoding="utf-8")

        print(
            f"{gen_file}: lines={len(result.lines)} macros={len(result.macros)} "
            f"excluded_gmax={result.excluded_gmax} excluded_cosmetic={result.excluded_cosmetic}"
        )

    print(f"TOTAL excluded_gmax={total_gmax} excluded_cosmetic={total_cosmetic}")

    if args.debug_parse:
        parser = Phase2SpeciesParser()
        print("PARSE SUMMARY")
        for gen_file, result in sorted(results.items()):
            parsed_records = parser.parse_species(result)
            print(f"{gen_file}: parsed_species={len(parsed_records)}")

        target_key = args.sample_species.strip()
        found: Optional[SpeciesRecord] = None
        for result in results.values():
            for record in parser.parse_species(result):
                if record.species_key == target_key:
                    found = record
                    break
            if found is not None:
                break

        if found is not None:
            print("SAMPLE")
            print(
                f"{found.species_key}: HP={found.base_hp} Atk={found.base_attack} Def={found.base_defense} "
                f"SpA={found.base_sp_attack} SpD={found.base_sp_defense} Spe={found.base_speed} "
                f"Types=({found.type1},{found.type2}) "
                f"Abilities=({found.ability1},{found.ability2},{found.hidden_ability})"
            )
        else:
            print(f"SAMPLE {target_key}: not found")

    print(f"Parsing customized repo: {repo_root}")
    custom_species = parse_repo_species(repo_root)
    print(f"Parsing original repo: {clean_repo_root}")
    original_species = parse_repo_species(clean_repo_root)

    merged_by_gen: Dict[str, List[MergedSpeciesRecord]] = {}
    for gen_file in GEN_FILES:
        merged_by_gen[gen_file] = merge_species(
            original_species.get(gen_file, []),
            custom_species.get(gen_file, []),
        )
        print(
            f"{gen_file}: original={len(original_species.get(gen_file, []))} "
            f"customized={len(custom_species.get(gen_file, []))} "
            f"merged={len(merged_by_gen[gen_file])}"
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    build_workbook(merged_by_gen, output_path)
    print(f"Spreadsheet written: {output_path}")


if __name__ == "__main__":
    main()
