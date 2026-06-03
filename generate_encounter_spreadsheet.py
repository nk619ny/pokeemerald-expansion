#!/usr/bin/env python3
"""
Generate a wild encounter spreadsheet (wild_encounters.xlsx) from pokeemerald-expansion data.

Reads:
  - src/data/wild_encounters.json       (encounter tables)
  - src/data/region_map/region_map_sections.json (MAPSEC display names)
  - include/constants/region_map_sections.h      (Kanto MAPSEC range)
  - data/maps/*/map.json                (map->MAPSEC, berry tree detection)

Requires: pip install openpyxl
"""

import json
import glob
import os
import re
import sys
from collections import defaultdict, OrderedDict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

# ---------------------------------------------------------------------------
# Paths (relative to repo root)
# ---------------------------------------------------------------------------
REPO_ROOT = os.path.dirname(os.path.abspath(__file__))
WILD_ENCOUNTERS_JSON = os.path.join(REPO_ROOT, "src", "data", "wild_encounters.json")
REGION_MAP_JSON = os.path.join(REPO_ROOT, "src", "data", "region_map", "region_map_sections.json")
REGION_MAP_H = os.path.join(REPO_ROOT, "include", "constants", "region_map_sections.h")
MAPS_DIR = os.path.join(REPO_ROOT, "data", "maps")
OUTPUT_FILE = os.path.join(REPO_ROOT, "wild_encounters.xlsx")

# ---------------------------------------------------------------------------
# Underwater MAPSEC -> Surface MAPSEC merge mapping
# ---------------------------------------------------------------------------
UNDERWATER_TO_SURFACE = {
    "MAPSEC_UNDERWATER_105": "MAPSEC_ROUTE_105",
    "MAPSEC_UNDERWATER_124": "MAPSEC_ROUTE_124",
    "MAPSEC_UNDERWATER_125": "MAPSEC_ROUTE_125",
    "MAPSEC_UNDERWATER_126": "MAPSEC_ROUTE_126",
    "MAPSEC_UNDERWATER_127": "MAPSEC_ROUTE_127",
    "MAPSEC_UNDERWATER_128": "MAPSEC_ROUTE_128",
    "MAPSEC_UNDERWATER_129": "MAPSEC_ROUTE_129",
    "MAPSEC_UNDERWATER_SOOTOPOLIS": "MAPSEC_SOOTOPOLIS_CITY",
    "MAPSEC_UNDERWATER_SEAFLOOR_CAVERN": "MAPSEC_SEAFLOOR_CAVERN",
    "MAPSEC_UNDERWATER_SEALED_CHAMBER": "MAPSEC_SEALED_CHAMBER",
    "MAPSEC_UNDERWATER_MARINE_CAVE": "MAPSEC_MARINE_CAVE",
}

# ---------------------------------------------------------------------------
# Aesthetic form grouping
# ---------------------------------------------------------------------------
AESTHETIC_BASES = {
    "FLABEBE", "FLOETTE", "FLORGES", "MINIOR", "ALCREMIE",
    "SHELLOS", "GASTRODON", "MAUSHOLD", "DUDUNSPARCE", "TATSUGIRI",
}

BATTLE_FORM_SUFFIXES = {"MEGA", "GMAX", "PRIMAL"}


def get_aesthetic_base(species_name):
    """
    If species_name (without SPECIES_ prefix) belongs to an aesthetic form group,
    return the base name. Otherwise return None.
    """
    parts = species_name.split("_")
    if parts[0] in AESTHETIC_BASES:
        for p in parts[1:]:
            if p in BATTLE_FORM_SUFFIXES:
                return None
        return parts[0]
    return None


# ---------------------------------------------------------------------------
# Species name formatting
# ---------------------------------------------------------------------------
SPECIAL_NAMES = {
    "NIDORAN_F": "Nidoran\u2640",
    "NIDORAN_M": "Nidoran\u2642",
    "MR_MIME": "Mr. Mime",
    "MR_MIME_GALAR": "Mr. Mime (Galar)",
    "MR_RIME": "Mr. Rime",
    "MIME_JR": "Mime Jr.",
    "HO_OH": "Ho-Oh",
    "PORYGON_Z": "Porygon-Z",
    "PORYGON2": "Porygon2",
    "JANGMO_O": "Jangmo-o",
    "HAKAMO_O": "Hakamo-o",
    "KOMMO_O": "Kommo-o",
    "KOMMO_O_TOTEM": "Kommo-o (Totem)",
    "TYPE_NULL": "Type: Null",
    "TAPU_KOKO": "Tapu Koko",
    "TAPU_LELE": "Tapu Lele",
    "TAPU_BULU": "Tapu Bulu",
    "TAPU_FINI": "Tapu Fini",
    "FLABEBE": "Flab\u00e9b\u00e9",
    "CHI_YU": "Chi-Yu",
    "CHIEN_PAO": "Chien-Pao",
    "WO_CHIEN": "Wo-Chien",
    "TING_LU": "Ting-Lu",
    "GREAT_TUSK": "Great Tusk",
    "SCREAM_TAIL": "Scream Tail",
    "BRUTE_BONNET": "Brute Bonnet",
    "FLUTTER_MANE": "Flutter Mane",
    "SLITHER_WING": "Slither Wing",
    "SANDY_SHOCKS": "Sandy Shocks",
    "IRON_TREADS": "Iron Treads",
    "IRON_BUNDLE": "Iron Bundle",
    "IRON_HANDS": "Iron Hands",
    "IRON_JUGULIS": "Iron Jugulis",
    "IRON_MOTH": "Iron Moth",
    "IRON_THORNS": "Iron Thorns",
    "IRON_VALIANT": "Iron Valiant",
    "IRON_LEAVES": "Iron Leaves",
    "IRON_BOULDER": "Iron Boulder",
    "IRON_CROWN": "Iron Crown",
    "ROARING_MOON": "Roaring Moon",
    "WALKING_WAKE": "Walking Wake",
    "GOUGING_FIRE": "Gouging Fire",
    "RAGING_BOLT": "Raging Bolt",
    "OGERPON_TEAL_MASK": "Ogerpon (Teal)",
    "OGERPON_WELLSPRING_MASK": "Ogerpon (Wellspring)",
    "OGERPON_HEARTHFLAME_MASK": "Ogerpon (Hearthflame)",
    "OGERPON_CORNERSTONE_MASK": "Ogerpon (Cornerstone)",
}

# Special display name overrides applied AFTER normal formatting
# These apply globally to any encounter table
DISPLAY_NAME_OVERRIDES = {
    "SPECIES_LYCANROC_MIDNIGHT": "Lycanroc Dusk / Midnight",
}

# Per-MAPSEC species display overrides: {mapsec: {species_raw: display_name}}
# Applied only to the rock_smash table for that MAPSEC
MAPSEC_ROCK_SMASH_OVERRIDES = {
    "MAPSEC_ROUTE_114": {
        "SPECIES_SOLROCK": "Solrock / Lunatone",
        "SPECIES_TYRUNT": "Tyrunt / Amaura",
    },
}

# Per-MAPSEC species display overrides applied to ALL table types for that MAPSEC
# {mapsec: {species_raw: display_name}}
MAPSEC_SPECIES_OVERRIDES = {
    "MAPSEC_CAVE_OF_ORIGIN": {
        "SPECIES_FLOETTE_ETERNAL": "Floette Eternal",
    },
}

FORM_SUFFIXES = {
    "ALOLA": "Alolan",
    "GALAR": "Galarian",
    "HISUI": "Hisuian",
    "PALDEA": "Paldean",
    "MEGA": "Mega",
    "MEGA_X": "Mega X",
    "MEGA_Y": "Mega Y",
    "GMAX": "Gigantamax",
    "PRIMAL": "Primal",
    "ORIGIN": "Origin",
    "THERIAN": "Therian",
    "TOTEM": "Totem",
    "CROWNED": "Crowned",
    "ETERNAMAX": "Eternamax",
    "BLACK": "Black",
    "WHITE": "White",
    "SHADOW": "Shadow",
    "DUSK_MANE": "Dusk Mane",
    "DAWN_WINGS": "Dawn Wings",
    "ULTRA": "Ultra",
    "ICE_RIDER": "Ice Rider",
    "SHADOW_RIDER": "Shadow Rider",
    "BLOODMOON": "Blood Moon",
    "AQUA_BREED": "Aqua Breed",
    "BLAZE_BREED": "Blaze Breed",
    "COMBAT_BREED": "Combat Breed",
}


def format_species_name(raw):
    """Convert SPECIES_XYZ to a display name."""
    # Check global overrides first
    if raw in DISPLAY_NAME_OVERRIDES:
        return DISPLAY_NAME_OVERRIDES[raw]

    name = raw.replace("SPECIES_", "")

    aesthetic_base = get_aesthetic_base(name)
    if aesthetic_base is not None:
        if aesthetic_base == "FLABEBE":
            return "Flab\u00e9b\u00e9"
        return aesthetic_base.capitalize()

    if name in SPECIAL_NAMES:
        return SPECIAL_NAMES[name]

    for suffix_key, suffix_display in sorted(FORM_SUFFIXES.items(), key=lambda x: -len(x[0])):
        if name.endswith("_" + suffix_key):
            base = name[: -(len(suffix_key) + 1)]
            base_formatted = format_base_name(base)
            return f"{base_formatted} ({suffix_display})"

    return format_base_name(name)


def format_base_name(name):
    """Format a base species name: Title case, handle special characters."""
    if name in SPECIAL_NAMES:
        return SPECIAL_NAMES[name]
    parts = name.split("_")
    formatted_parts = []
    for part in parts:
        if part:
            formatted_parts.append(part.capitalize())
    return " ".join(formatted_parts)


def get_grouping_key(raw):
    """Return the key used to group/aggregate a species."""
    name = raw.replace("SPECIES_", "")
    aesthetic_base = get_aesthetic_base(name)
    if aesthetic_base is not None:
        return "SPECIES_" + aesthetic_base
    return raw


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------
def load_mapsec_names():
    """Load MAPSEC id -> display name from region_map_sections.json."""
    with open(REGION_MAP_JSON, "r", encoding="utf-8") as f:
        data = json.load(f)
    mapping = {}
    ordered_ids = []
    for entry in data["map_sections"]:
        mapsec_id = entry["id"]
        display_name = entry.get("name", mapsec_id.replace("MAPSEC_", "").replace("_", " "))
        mapping[mapsec_id] = display_name
        ordered_ids.append(mapsec_id)
    return mapping, ordered_ids


def load_kanto_mapsecs(ordered_ids):
    """Determine the set of Kanto/FRLG MAPSECs to exclude."""
    try:
        start_idx = ordered_ids.index("MAPSEC_PALLET_TOWN")
        end_idx = ordered_ids.index("MAPSEC_SPECIAL_AREA")
        return set(ordered_ids[start_idx : end_idx + 1])
    except ValueError:
        kanto_set = set()
        with open(REGION_MAP_H, "r", encoding="utf-8") as f:
            content = f.read()
        in_kanto = False
        for line in content.splitlines():
            line = line.strip().rstrip(",")
            if "MAPSEC_PALLET_TOWN" in line:
                in_kanto = True
            if in_kanto:
                m = re.match(r"(MAPSEC_\w+)", line)
                if m:
                    kanto_set.add(m.group(1))
            if "MAPSEC_SPECIAL_AREA" in line:
                break
        return kanto_set


def load_map_data(kanto_mapsecs):
    """Scan all map.json files. Returns map_id_to_info dict. Excludes FRLG maps."""
    map_id_to_info = {}
    map_json_pattern = os.path.join(MAPS_DIR, "*", "map.json")
    for path in glob.glob(map_json_pattern):
        folder_name = os.path.basename(os.path.dirname(path))
        if folder_name.endswith("_Frlg"):
            continue
        with open(path, "r", encoding="utf-8") as f:
            try:
                data = json.load(f)
            except json.JSONDecodeError:
                continue
        map_id = data.get("id", "")
        mapsec = data.get("region_map_section", "")
        map_name = data.get("name", folder_name)
        if mapsec in kanto_mapsecs:
            continue
        if data.get("region") == "REGION_KANTO":
            continue
        has_berry = False
        for obj in data.get("object_events", []):
            if obj.get("graphics_id") == "OBJ_EVENT_GFX_BERRY_TREE":
                has_berry = True
                break
        map_id_to_info[map_id] = {
            "mapsec": mapsec,
            "name": map_name,
            "has_berry_tree": has_berry,
        }
    return map_id_to_info


def load_encounters():
    """Load wild_encounters.json and return encounter rates + encounter entries."""
    with open(WILD_ENCOUNTERS_JSON, "r", encoding="utf-8") as f:
        data = json.load(f)
    main_group = None
    for group in data["wild_encounter_groups"]:
        if group.get("for_maps") is True and group["label"] == "gWildMonHeaders":
            main_group = group
            break
    if main_group is None:
        print("ERROR: Could not find gWildMonHeaders encounter group")
        sys.exit(1)
    rates = {}
    fishing_groups = {}
    for field in main_group["fields"]:
        ftype = field["type"]
        rates[ftype] = field["encounter_rates"]
        if ftype == "fishing_mons" and "groups" in field:
            fishing_groups = field["groups"]
    return rates, fishing_groups, main_group["encounters"]


# ---------------------------------------------------------------------------
# Encounter processing   (Night = Day duplication, underwater merge, overrides)
# ---------------------------------------------------------------------------
def aggregate_species(slots, rates, rock_smash_overrides=None, overrides=None):
    """Aggregate encounter slots into {display_name: probability}."""
    total_rate = sum(rates)
    if total_rate == 0:
        return {}
    species_prob = {}
    for slot, rate_val in zip(slots, rates):
        sp = slot["species"]
        display = format_species_name(sp)
        if overrides and sp in overrides:
            display = overrides[sp]
        if rock_smash_overrides and sp in rock_smash_overrides:
            display = rock_smash_overrides[sp]
        species_prob[display] = species_prob.get(display, 0.0) + rate_val / total_rate
    # Merge aesthetic forms
    grouped = {}
    for display, prob in species_prob.items():
        key = get_grouping_key(display)
        grouped[key] = grouped.get(key, 0.0) + prob
    return grouped


def process_encounters(encounters, rates, fishing_groups, map_id_to_info,
                       mapsec_names, kanto_mapsecs):
    """
    Build mapsec_data: {mapsec: {map_name: {encounter_type: {species: prob}}}}.

    The JSON has separate encounter entries for Day and Night per map,
    distinguished by base_label ending in _Day or _Night.
    - Land/Surf Day subtables use the _Day entry.
    - Land/Surf Night subtables use the _Night entry.
    - Fishing and Rock Smash always use the _Day entry.
    - If no _Night entry exists for Land/Surf but _Day does, duplicate Day→Night.

    Also applies:
      - Underwater MAPSEC merging into surface MAPSEC
      - Rock smash display name overrides per MAPSEC
    """
    fishing_rate_keys = ["old_rod", "good_rod", "super_rod"]
    if fishing_groups:
        fishing_rate_keys = list(fishing_groups.keys())

    # Group encounter entries by map_id into day/night
    map_encounters = defaultdict(dict)
    for entry in encounters:
        map_id = entry.get("map", "")
        if map_id not in map_id_to_info:
            continue
        label = entry.get("base_label", "")
        if label.endswith("_Night"):
            map_encounters[map_id]["night"] = entry
        elif label.endswith("_Day"):
            map_encounters[map_id]["day"] = entry
        # Skip _Evening entries

    mapsec_data = {}

    for map_id, entries in map_encounters.items():
        info = map_id_to_info[map_id]
        mapsec = info["mapsec"]
        map_name = info["name"]
        if mapsec in kanto_mapsecs:
            continue

        day_entry = entries.get("day", {})
        night_entry = entries.get("night", {})

        # Check if this map has any encounter data at all
        has_any = False
        for enc_type in ["land_mons", "water_mons", "rock_smash_mons", "fishing_mons"]:
            if enc_type in day_entry or enc_type in night_entry:
                has_any = True
                break
        if not has_any:
            continue

        # Determine overrides for this MAPSEC
        rs_overrides = MAPSEC_ROCK_SMASH_OVERRIDES.get(mapsec)
        ms_overrides = MAPSEC_SPECIES_OVERRIDES.get(mapsec)

        encounter_dict = {}
        land_rate = rates.get("land_mons", [])
        water_rate = rates.get("water_mons", [])

        # --- Land (Day from _Day entry, Night from _Night entry) ---
        day_land = day_entry.get("land_mons", {}).get("mons", [])
        night_land = night_entry.get("land_mons", {}).get("mons", [])
        land_day_agg = aggregate_species(day_land, land_rate, overrides=ms_overrides) if day_land else {}
        land_night_agg = aggregate_species(night_land, land_rate, overrides=ms_overrides) if night_land else {}
        if not land_night_agg and land_day_agg:
            land_night_agg = dict(land_day_agg)
        if land_day_agg:
            encounter_dict["land_day"] = land_day_agg
        if land_night_agg:
            encounter_dict["land_night"] = land_night_agg

        # --- Surf (Day from _Day entry, Night from _Night entry) ---
        day_water = day_entry.get("water_mons", {}).get("mons", [])
        night_water = night_entry.get("water_mons", {}).get("mons", [])
        surf_day_agg = aggregate_species(day_water, water_rate, overrides=ms_overrides) if day_water else {}
        surf_night_agg = aggregate_species(night_water, water_rate, overrides=ms_overrides) if night_water else {}
        if not surf_night_agg and surf_day_agg:
            surf_night_agg = dict(surf_day_agg)
        if surf_day_agg:
            encounter_dict["surf_day"] = surf_day_agg
        if surf_night_agg:
            encounter_dict["surf_night"] = surf_night_agg

        # --- Rock Smash (always from _Day entry) ---
        rs_rate = rates.get("rock_smash_mons", [])
        day_rock = day_entry.get("rock_smash_mons", {}).get("mons", [])
        rs_agg = aggregate_species(day_rock, rs_rate, rs_overrides, overrides=ms_overrides) if day_rock else {}
        if rs_agg:
            encounter_dict["rock_smash"] = rs_agg

        # --- Fishing (always from _Day entry) ---
        fish_rate = rates.get("fishing_mons", [])
        day_fish = day_entry.get("fishing_mons", {}).get("mons", [])
        if day_fish and fish_rate:
            group_size = 10
            for gi, group_key in enumerate(fishing_rate_keys):
                start = gi * group_size
                end = start + group_size
                sub_slots = day_fish[start:end]
                sub_rates = fish_rate[start:end]
                if sub_slots:
                    agg = aggregate_species(sub_slots, sub_rates, overrides=ms_overrides)
                    if agg:
                        encounter_dict["fish_" + group_key] = agg

        if encounter_dict:
            if mapsec not in mapsec_data:
                mapsec_data[mapsec] = {}
            mapsec_data[mapsec][map_name] = encounter_dict

    # --- Merge underwater into surface ---
    for uw_mapsec, surface_mapsec in UNDERWATER_TO_SURFACE.items():
        if uw_mapsec in mapsec_data:
            if surface_mapsec not in mapsec_data:
                mapsec_data[surface_mapsec] = {}
            for map_name, enc_dict in mapsec_data[uw_mapsec].items():
                uw_label = map_name + " (Underwater)"
                mapsec_data[surface_mapsec][uw_label] = enc_dict
            del mapsec_data[uw_mapsec]

    return mapsec_data



# ---------------------------------------------------------------------------
# Color definitions (per-subtable pastel fills)
# ---------------------------------------------------------------------------
# Table group headers
LAND_HEADER_FILL = PatternFill(start_color="7A9A3A", end_color="7A9A3A", fill_type="solid")
ROCK_HEADER_FILL = PatternFill(start_color="A0845C", end_color="A0845C", fill_type="solid")
FISH_HEADER_FILL = PatternFill(start_color="2E8B8B", end_color="2E8B8B", fill_type="solid")
SURF_HEADER_FILL = PatternFill(start_color="4A7FAF", end_color="4A7FAF", fill_type="solid")

# Sub-headers
LAND_DAY_SUBHEADER = PatternFill(start_color="B5CC6A", end_color="B5CC6A", fill_type="solid")
LAND_NIGHT_SUBHEADER = PatternFill(start_color="5A7D2A", end_color="5A7D2A", fill_type="solid")
ROCK_SUBHEADER = PatternFill(start_color="C4A97D", end_color="C4A97D", fill_type="solid")
OLD_ROD_SUBHEADER = PatternFill(start_color="66CDAA", end_color="66CDAA", fill_type="solid")
GOOD_ROD_SUBHEADER = PatternFill(start_color="48B0A0", end_color="48B0A0", fill_type="solid")
SUPER_ROD_SUBHEADER = PatternFill(start_color="2E8B7A", end_color="2E8B7A", fill_type="solid")
SURF_DAY_SUBHEADER = PatternFill(start_color="7AB8D9", end_color="7AB8D9", fill_type="solid")
SURF_NIGHT_SUBHEADER = PatternFill(start_color="3A6B8F", end_color="3A6B8F", fill_type="solid")

# Data cell fills
LAND_DAY_FILL = PatternFill(start_color="E8F5C8", end_color="E8F5C8", fill_type="solid")
LAND_NIGHT_FILL = PatternFill(start_color="C8DFA0", end_color="C8DFA0", fill_type="solid")
ROCK_FILL = PatternFill(start_color="F0E0C8", end_color="F0E0C8", fill_type="solid")
OLD_ROD_FILL = PatternFill(start_color="D4F5EC", end_color="D4F5EC", fill_type="solid")
GOOD_ROD_FILL = PatternFill(start_color="BDE8DC", end_color="BDE8DC", fill_type="solid")
SUPER_ROD_FILL = PatternFill(start_color="A6DDD0", end_color="A6DDD0", fill_type="solid")
SURF_DAY_FILL = PatternFill(start_color="D6EEFC", end_color="D6EEFC", fill_type="solid")
SURF_NIGHT_FILL = PatternFill(start_color="B0D4EE", end_color="B0D4EE", fill_type="solid")

MAP_NAME_FILL = PatternFill(start_color="3C3C3C", end_color="3C3C3C", fill_type="solid")

# Fonts
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
SUBHEADER_FONT_DARK = Font(bold=True, color="1A1A1A", size=10)
MAP_NAME_FONT = Font(bold=True, color="FFFFFF", size=12)
DATA_FONT = Font(size=10)
NOTE_FONT = Font(italic=True, size=10, color="2E7D32")
PCT_FONT = Font(size=10, color="555555")

THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
LEFT_ALIGN = Alignment(horizontal="left", vertical="center")
WRAP_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)


# ---------------------------------------------------------------------------
# Dynamic column layout (with gap columns between table groups)
# ---------------------------------------------------------------------------
# Table group definitions in display order: Land -> Rock Smash -> Fishing -> Surf
# Each: (group_key, header_title, header_fill,
#         [(data_key, sub_label, sub_header_fill, sub_header_font, data_fill), ...])
TABLE_DEFS = [
    ("land", "Land Encounters", LAND_HEADER_FILL, [
        ("land_day", "Day", LAND_DAY_SUBHEADER, SUBHEADER_FONT_DARK, LAND_DAY_FILL),
        ("land_night", "Night", LAND_NIGHT_SUBHEADER, SUBHEADER_FONT, LAND_NIGHT_FILL),
    ]),
    ("rock", "Cut Tree / Rock Smash", ROCK_HEADER_FILL, [
        ("rock_smash", "", ROCK_SUBHEADER, SUBHEADER_FONT, ROCK_FILL),
    ]),
    ("fish", "Fishing Encounters", FISH_HEADER_FILL, [
        ("fish_old_rod", "Old Rod", OLD_ROD_SUBHEADER, SUBHEADER_FONT_DARK, OLD_ROD_FILL),
        ("fish_good_rod", "Good Rod", GOOD_ROD_SUBHEADER, SUBHEADER_FONT, GOOD_ROD_FILL),
        ("fish_super_rod", "Super Rod", SUPER_ROD_SUBHEADER, SUBHEADER_FONT, SUPER_ROD_FILL),
    ]),
    ("surf", "Surf Encounters", SURF_HEADER_FILL, [
        ("surf_day", "Day", SURF_DAY_SUBHEADER, SUBHEADER_FONT_DARK, SURF_DAY_FILL),
        ("surf_night", "Night", SURF_NIGHT_SUBHEADER, SUBHEADER_FONT, SURF_NIGHT_FILL),
    ]),
]


def compute_layout(mapsec_maps):
    """
    Determine which table groups are present across all maps in a MAPSEC.
    Returns (layout, note_col, last_col).
    layout: [(header_title, header_fill, start_col, end_col, sub_info)]
    sub_info: [(data_key, sub_label, sub_header_fill, sub_header_font,
                data_fill, name_col, pct_col)]
    """
    present_keys = set()
    for enc_dict in mapsec_maps.values():
        present_keys.update(enc_dict.keys())

    layout = []
    col = 1
    first_table = True

    for _gkey, header_title, header_fill, sub_defs in TABLE_DEFS:
        active = [(dk, sl, shf, shfont, df)
                  for dk, sl, shf, shfont, df in sub_defs
                  if dk in present_keys]
        if not active:
            continue

        if not first_table:
            col += 1  # gap column

        start = col
        sub_info = []
        for dk, sl, shf, shfont, df in active:
            sub_info.append((dk, sl, shf, shfont, df, col, col + 1))
            col += 2
        end = col - 1

        layout.append((header_title, header_fill, start, end, sub_info))
        first_table = False

    note_col = (col + 1) if layout else 1
    last_col = note_col
    return layout, note_col, last_col


# ---------------------------------------------------------------------------
# Spreadsheet writing helpers
# ---------------------------------------------------------------------------
def write_subtable_data(ws, start_row, col_name, col_pct, data_dict, fill, max_rows):
    """Write species/percentage data into a pair of columns with uniform fill."""
    sorted_species = sorted(data_dict.items(), key=lambda x: (-x[1], x[0]))
    for i, (species, prob) in enumerate(sorted_species):
        row = start_row + i
        cell_name = ws.cell(row=row, column=col_name, value=species)
        cell_name.font = DATA_FONT
        cell_name.fill = fill
        cell_name.border = THIN_BORDER
        cell_name.alignment = LEFT_ALIGN

        cell_pct = ws.cell(row=row, column=col_pct, value=f"{prob * 100:.0f}%")
        cell_pct.font = PCT_FONT
        cell_pct.fill = fill
        cell_pct.border = THIN_BORDER
        cell_pct.alignment = CENTER_ALIGN

    # Fill remaining empty rows with background color for consistent look
    for i in range(len(sorted_species), max_rows):
        row = start_row + i
        for col in (col_name, col_pct):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.border = THIN_BORDER


def write_map_section(ws, row, map_name, encounter_dict, layout, note_col,
                      last_col, map_id_to_info):
    """
    Write a single map's encounter section (map name, headers, sub-headers, data).
    Returns the next available row after this section.
    """
    # Check which table groups this map actually has
    map_keys = set(encounter_dict.keys())

    # Compute max data rows across all present sub-columns
    max_data_rows = 0
    for _htitle, _hfill, _sc, _ec, sub_info in layout:
        for dk, _sl, _shf, _shfont, _df, _nc, _pc in sub_info:
            max_data_rows = max(max_data_rows, len(encounter_dict.get(dk, {})))

    if max_data_rows == 0:
        return row

    # ---- Row 1: Map name bar ----
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    cell = ws.cell(row=row, column=1, value=map_name.replace("_", " "))
    cell.font = MAP_NAME_FONT
    cell.fill = MAP_NAME_FILL
    cell.alignment = CENTER_ALIGN
    row += 1

    # Determine which layout groups this map has data for
    def map_has_group(sub_info):
        return any(dk in map_keys for dk, *_ in sub_info)

    # ---- Row 2: Table title headers (per-map, only for present groups) ----
    title_row = row
    for htitle, hfill, sc, ec, sub_info in layout:
        if not map_has_group(sub_info):
            continue
        ws.merge_cells(start_row=title_row, start_column=sc, end_row=title_row, end_column=ec)
        cell = ws.cell(row=title_row, column=sc, value=htitle)
        cell.font = HEADER_FONT
        cell.fill = hfill
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        for c in range(sc, ec + 1):
            ws.cell(row=title_row, column=c).fill = hfill
            ws.cell(row=title_row, column=c).border = THIN_BORDER
    row += 1

    # ---- Row 3: Sub-headers (per-map) ----
    subheader_row = row
    for _htitle, _hfill, _sc, _ec, sub_info in layout:
        if not map_has_group(sub_info):
            continue
        for dk, slabel, shfill, shfont, _df, nc, pc in sub_info:
            ws.merge_cells(start_row=subheader_row, start_column=nc,
                           end_row=subheader_row, end_column=pc)
            cell = ws.cell(row=subheader_row, column=nc, value=slabel)
            cell.font = shfont
            cell.fill = shfill
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
            ws.cell(row=subheader_row, column=pc).fill = shfill
            ws.cell(row=subheader_row, column=pc).border = THIN_BORDER
    row += 1

    # ---- Data rows ----
    data_start = row
    for _htitle, _hfill, _sc, _ec, sub_info in layout:
        for dk, _sl, _shf, _shfont, dfill, nc, pc in sub_info:
            data = encounter_dict.get(dk, {})
            if data or map_has_group(sub_info):
                write_subtable_data(ws, data_start, nc, pc, data, dfill, max_data_rows)

    # ---- Note column ----
    has_berry = False
    clean_name = map_name.replace(" (Underwater)", "")
    for _mid, minfo in map_id_to_info.items():
        if minfo["name"] == clean_name and minfo.get("has_berry_tree"):
            has_berry = True
            break
    if has_berry:
        cell = ws.cell(row=data_start, column=note_col, value="Berry Tree Encounters available.")
        cell.font = NOTE_FONT
        cell.alignment = WRAP_ALIGN

    row = data_start + max_data_rows + 1  # +1 for blank separator row
    return row


def set_column_widths(ws, layout, note_col):
    """Set column widths dynamically based on layout, including gap columns."""
    name_width = 18
    pct_width = 7
    gap_width = 2

    for _htitle, _hfill, _sc, _ec, sub_info in layout:
        for _dk, _sl, _shf, _shfont, _df, nc, pc in sub_info:
            ws.column_dimensions[get_column_letter(nc)].width = name_width
            ws.column_dimensions[get_column_letter(pc)].width = pct_width

    # Set gap columns
    gap_cols = set()
    prev_end = 0
    for _htitle, _hfill, sc, ec, _sub in layout:
        if prev_end > 0 and sc > prev_end + 1:
            gap_cols.add(prev_end + 1)
        prev_end = ec
    for gc in gap_cols:
        ws.column_dimensions[get_column_letter(gc)].width = gap_width

    # Gap before note column
    if layout:
        last_end = layout[-1][3]
        if note_col > last_end + 1:
            ws.column_dimensions[get_column_letter(last_end + 1)].width = gap_width

    ws.column_dimensions[get_column_letter(note_col)].width = 30


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    print("Loading MAPSEC data...")
    mapsec_names, ordered_ids = load_mapsec_names()
    kanto_mapsecs = load_kanto_mapsecs(ordered_ids)
    print(f"  {len(kanto_mapsecs)} Kanto MAPSECs to exclude")

    print("Loading map data...")
    map_id_to_info = load_map_data(kanto_mapsecs)
    print(f"  {len(map_id_to_info)} Emerald maps loaded")

    print("Loading encounters...")
    rates, fishing_groups, encounters = load_encounters()
    print(f"  {len(encounters)} encounter entries loaded")

    print("Processing encounters...")
    mapsec_data = process_encounters(
        encounters, rates, fishing_groups, map_id_to_info,
        mapsec_names, kanto_mapsecs,
    )
    print(f"  {len(mapsec_data)} MAPSECs with encounters")

    print("Generating spreadsheet...")
    wb = Workbook()
    wb.remove(wb.active)

    # Order MAPSECs by their position in the enum
    mapsec_order = {m: i for i, m in enumerate(ordered_ids)}
    sorted_mapsecs = sorted(mapsec_data.keys(), key=lambda m: mapsec_order.get(m, 9999))

    tab_names_used = set()
    for mapsec in sorted_mapsecs:
        maps = mapsec_data[mapsec]
        if not maps:
            continue

        display_name = mapsec_names.get(mapsec, mapsec.replace("MAPSEC_", "").replace("_", " "))

        # Excel tab names max 31 chars, no special chars
        tab_name = display_name[:31]
        for ch in ["\\", "/", "*", "?", ":", "[", "]"]:
            tab_name = tab_name.replace(ch, "")
        tab_name = tab_name.strip()
        if not tab_name:
            tab_name = mapsec[:31]
        base_tab = tab_name
        counter = 2
        while tab_name in tab_names_used:
            suffix = f" ({counter})"
            tab_name = base_tab[: 31 - len(suffix)] + suffix
            counter += 1
        tab_names_used.add(tab_name)

        ws = wb.create_sheet(title=tab_name)

        # Compute dynamic layout for this MAPSEC
        layout, note_col, last_col = compute_layout(maps)

        if not layout:
            row = 1
            for map_name in sorted(maps.keys()):
                ws.cell(row=row, column=1, value=map_name.replace("_", " "))
                row += 1
            continue

        set_column_widths(ws, layout, note_col)

        # Freeze top row
        ws.freeze_panes = "A2"

        current_row = 1
        sorted_maps = sorted(maps.keys())
        for map_name in sorted_maps:
            encounter_dict = maps[map_name]
            current_row = write_map_section(
                ws, current_row, map_name, encounter_dict,
                layout, note_col, last_col, map_id_to_info,
            )

    if not wb.sheetnames:
        print("WARNING: No encounter data found. Creating empty workbook.")
        wb.create_sheet(title="No Data")

    wb.save(OUTPUT_FILE)
    print(f"Done! Saved to {OUTPUT_FILE}")
    print(f"  {len(wb.sheetnames)} tabs created")


if __name__ == "__main__":
    main()
